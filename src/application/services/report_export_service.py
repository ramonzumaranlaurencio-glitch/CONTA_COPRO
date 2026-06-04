from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle


class ReportExportService:
    def build_financial_pack_xlsx(self, payload: dict[str, Any]) -> bytes:
        period = payload.get("period", "N/A")
        comparison = payload.get("comparison") or {}
        compare_period = (
            comparison.get("balance_sheet", {}).get("period")
            or comparison.get("income_statement", {}).get("period")
            or comparison.get("cash_flow", {}).get("period")
            or "Comparado"
        )

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_balance = wb.create_sheet("Balance")
        ws_income = wb.create_sheet("Resultados")
        ws_cash = wb.create_sheet("Flujo")
        ws_ratios = wb.create_sheet("Ratios")

        self._fill_summary_sheet(ws_summary, payload, period, compare_period)
        self._fill_balance_sheet(ws_balance, payload, period, compare_period)
        self._fill_income_sheet(ws_income, payload, period, compare_period)
        self._fill_cash_sheet(ws_cash, payload, period, compare_period)
        self._fill_ratios_sheet(ws_ratios, payload)

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def build_financial_pack_pdf(self, payload: dict[str, Any]) -> bytes:
        period = payload.get("period", "N/A")
        comparison = payload.get("comparison") or {}
        compare_period = (
            comparison.get("balance_sheet", {}).get("period")
            or comparison.get("income_statement", {}).get("period")
            or comparison.get("cash_flow", {}).get("period")
            or "Comparado"
        )

        income = payload.get("income_statement") or {}
        income_cmp = comparison.get("income_statement") or {}
        balance = payload.get("balance_sheet") or {}
        balance_cmp = comparison.get("balance_sheet") or {}
        cash = payload.get("cash_flow") or {}
        cash_cmp = comparison.get("cash_flow") or {}
        ratios = payload.get("ratios") or {}

        stream = BytesIO()
        doc = SimpleDocTemplate(stream, pagesize=A4, title="Reporte Financiero")
        styles = getSampleStyleSheet()
        content = [
            Paragraph("CONTA_PRO Enterprise - Reporte Financiero", styles["Title"]),
            Paragraph(f"Periodo actual: {period} | Periodo comparado: {compare_period}", styles["Normal"]),
            Spacer(1, 12),
            Paragraph("Estado de Resultados", styles["Heading3"]),
            self._pdf_table(
                ["Indicador", period, compare_period],
                [
                    ["Ingresos", self._safe_money(income.get("revenue")), self._safe_money(income_cmp.get("revenue"))],
                    ["Costos", self._safe_money(income.get("cost")), self._safe_money(income_cmp.get("cost"))],
                    ["Gastos", self._safe_money(income.get("expenses")), self._safe_money(income_cmp.get("expenses"))],
                    ["Utilidad Operativa", self._safe_money(income.get("operating_profit")), self._safe_money(income_cmp.get("operating_profit"))],
                ],
            ),
            Spacer(1, 10),
            Paragraph("Balance General", styles["Heading3"]),
            self._pdf_table(
                ["Indicador", period, compare_period],
                [
                    ["Activos", self._safe_money(balance.get("assets")), self._safe_money(balance_cmp.get("assets"))],
                    ["Pasivos", self._safe_money(balance.get("liabilities")), self._safe_money(balance_cmp.get("liabilities"))],
                    ["Patrimonio", self._safe_money(balance.get("equity")), self._safe_money(balance_cmp.get("equity"))],
                    ["Check", self._safe_money(balance.get("check")), self._safe_money(balance_cmp.get("check"))],
                ],
            ),
            Spacer(1, 10),
            Paragraph("Flujo de Caja", styles["Heading3"]),
            self._pdf_table(
                ["Indicador", period, compare_period],
                [
                    ["Apertura", self._safe_money(cash.get("opening_cash")), self._safe_money(cash_cmp.get("opening_cash"))],
                    ["Movimiento Neto", self._safe_money(cash.get("net_cash_movement")), self._safe_money(cash_cmp.get("net_cash_movement"))],
                    ["Cierre", self._safe_money(cash.get("ending_cash")), self._safe_money(cash_cmp.get("ending_cash"))],
                ],
            ),
            Spacer(1, 10),
            Paragraph("Ratios", styles["Heading3"]),
            self._pdf_table(
                ["Ratio", "Valor"],
                [
                    ["Margen Operativo", self._safe_money(ratios.get("operating_margin"))],
                    ["Deuda / Patrimonio", self._safe_money(ratios.get("debt_to_equity"))],
                    ["Apalancamiento", self._safe_money(ratios.get("financial_leverage"))],
                ],
            ),
        ]

        doc.build(content)
        return stream.getvalue()

    def _fill_summary_sheet(self, ws, payload: dict[str, Any], period: str, compare_period: str) -> None:
        income = payload.get("income_statement") or {}
        income_cmp = (payload.get("comparison") or {}).get("income_statement") or {}

        ws["A1"] = "Reporte Financiero Comparativo"
        ws["A2"] = f"Actual: {period}"
        ws["B2"] = f"Comparado: {compare_period}"
        ws["A4"] = "Indicador"
        ws["B4"] = period
        ws["C4"] = compare_period

        summary_rows = [
            ("Ingresos", self._safe_float(income.get("revenue")), self._safe_float(income_cmp.get("revenue"))),
            ("Costos", self._safe_float(income.get("cost")), self._safe_float(income_cmp.get("cost"))),
            ("Gastos", self._safe_float(income.get("expenses")), self._safe_float(income_cmp.get("expenses"))),
            ("Utilidad Operativa", self._safe_float(income.get("operating_profit")), self._safe_float(income_cmp.get("operating_profit"))),
        ]
        for idx, row in enumerate(summary_rows, start=5):
            ws.cell(row=idx, column=1, value=row[0])
            ws.cell(row=idx, column=2, value=row[1])
            ws.cell(row=idx, column=3, value=row[2])

        table = Table(displayName="ResumenTabla", ref="A4:C8")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

        chart = BarChart()
        chart.title = "Comparativo Estado de Resultados"
        chart.y_axis.title = "COP"
        chart.x_axis.title = "Indicador"
        data = Reference(ws, min_col=2, max_col=3, min_row=4, max_row=8)
        cats = Reference(ws, min_col=1, min_row=5, max_row=8)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "E4")

        self._style_sheet(ws)

    def _fill_balance_sheet(self, ws, payload: dict[str, Any], period: str, compare_period: str) -> None:
        balance = payload.get("balance_sheet") or {}
        balance_cmp = (payload.get("comparison") or {}).get("balance_sheet") or {}

        ws.append(["Cuenta", period, compare_period])
        ws.append(["Activos", self._safe_float(balance.get("assets")), self._safe_float(balance_cmp.get("assets"))])
        ws.append(["Pasivos", self._safe_float(balance.get("liabilities")), self._safe_float(balance_cmp.get("liabilities"))])
        ws.append(["Patrimonio", self._safe_float(balance.get("equity")), self._safe_float(balance_cmp.get("equity"))])
        ws.append(["Check", self._safe_float(balance.get("check")), self._safe_float(balance_cmp.get("check"))])

        table = Table(displayName="BalanceTabla", ref="A1:C5")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
        self._style_sheet(ws)

    def _fill_income_sheet(self, ws, payload: dict[str, Any], period: str, compare_period: str) -> None:
        income = payload.get("income_statement") or {}
        income_cmp = (payload.get("comparison") or {}).get("income_statement") or {}

        ws.append(["Cuenta", period, compare_period])
        ws.append(["Ingresos", self._safe_float(income.get("revenue")), self._safe_float(income_cmp.get("revenue"))])
        ws.append(["Costos", self._safe_float(income.get("cost")), self._safe_float(income_cmp.get("cost"))])
        ws.append(["Gastos", self._safe_float(income.get("expenses")), self._safe_float(income_cmp.get("expenses"))])
        ws.append(["Utilidad Bruta", self._safe_float(income.get("gross_profit")), self._safe_float(income_cmp.get("gross_profit"))])
        ws.append(["Utilidad Operativa", self._safe_float(income.get("operating_profit")), self._safe_float(income_cmp.get("operating_profit"))])

        table = Table(displayName="ResultadosTabla", ref="A1:C6")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws.add_table(table)
        self._style_sheet(ws)

    def _fill_cash_sheet(self, ws, payload: dict[str, Any], period: str, compare_period: str) -> None:
        cash = payload.get("cash_flow") or {}
        cash_cmp = (payload.get("comparison") or {}).get("cash_flow") or {}

        ws.append(["Indicador", period, compare_period])
        ws.append(["Apertura", self._safe_float(cash.get("opening_cash")), self._safe_float(cash_cmp.get("opening_cash"))])
        ws.append(["Movimiento Neto", self._safe_float(cash.get("net_cash_movement")), self._safe_float(cash_cmp.get("net_cash_movement"))])
        ws.append(["Cierre", self._safe_float(cash.get("ending_cash")), self._safe_float(cash_cmp.get("ending_cash"))])

        table = Table(displayName="FlujoTabla", ref="A1:C4")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium6", showRowStripes=True)
        ws.add_table(table)
        self._style_sheet(ws)

    def _fill_ratios_sheet(self, ws, payload: dict[str, Any]) -> None:
        ratios = payload.get("ratios") or {}
        ws.append(["Ratio", "Valor"])
        ws.append(["Margen Operativo", self._safe_float(ratios.get("operating_margin"))])
        ws.append(["Deuda / Patrimonio", self._safe_float(ratios.get("debt_to_equity"))])
        ws.append(["Apalancamiento", self._safe_float(ratios.get("financial_leverage"))])

        table = Table(displayName="RatiosTabla", ref="A1:B4")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
        ws.add_table(table)
        self._style_sheet(ws)

    def _style_sheet(self, ws) -> None:
        header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
            for cell in row:
                if cell.value is None:
                    continue
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value is None:
                    continue
                cell.border = border
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max(14, max_len + 2), 40)

    def _pdf_table(self, header: list[str], rows: list[list[str]]) -> PdfTable:
        table = PdfTable([header, *rows], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ]
            )
        )
        return table

    @staticmethod
    def _safe_float(value: Any) -> float:
        try:
            return float(value)
        except Exception:
            return 0.0

    @staticmethod
    def _safe_money(value: Any) -> str:
        try:
            return f"{float(value):,.2f}"
        except Exception:
            return "0.00"
